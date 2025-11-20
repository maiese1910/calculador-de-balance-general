
import pandas as pd
import os
from src.processor import export_balance_to_excel

# Create mock data
data = {
    'Cuenta': ['1101 - CAJA', '1102 - BANCOS'],
    'Debe': [15000.50, 0],
    'Haber': [0, 5000.25],
    'Saldo': [15000.50, -5000.25]
}
df = pd.DataFrame(data)

output_file = 'test_balance_export.xlsx'

print(f"Exporting to {output_file}...")
try:
    export_balance_to_excel(df, output_file)
    print("Export successful.")
    
    # Verify file exists
    if os.path.exists(output_file):
        print("File created.")
        # We could use openpyxl to verify formatting, but existence and no error is a good first step.
        import openpyxl
        wb = openpyxl.load_workbook(output_file)
        ws = wb['Balance']
        print(f"Sheet 'Balance' found. Dimensions: {ws.dimensions}")
        print(f"Cell B2 format: {ws['B2'].number_format}")
        if ws['B2'].number_format == '#,##0.00':
            print("TEST PASSED: Number format correct.")
        else:
            print(f"TEST FAILED: Number format is {ws['B2'].number_format}")
            
        # Clean up
        wb.close()
        os.remove(output_file)
        print("Test file cleaned up.")
    else:
        print("TEST FAILED: File not created.")
        
except Exception as e:
    print(f"TEST FAILED: Exception occurred: {e}")
